import openpyxl as xl
import os

class PicAndUnempMatcher:

    def __init__(self, rootDir, picsDir, unempData, MSAName, timePeriod):
        self.picsDir = picsDir + "/periodAvgs"
        self.unempData = unempData
        self.MSAName = MSAName
        self.timePeriod = timePeriod
        self.rootDir = rootDir
        #self.MSADir = rootDir + "/" + MSAName

    #Match picture with unemployment datum in Excel sheet
    def match(self):
        wb = xl.load_workbook(self.unempData, data_only = True)
        sheet = wb.active
        print("Unemployment data loaded")
        #Iterate through each row, looking for our MSA
        for i in range(1, sheet.max_row + 1):
            #NOTE: Row/column values come from specific unempData format -- not general
            if(sheet.cell(row=i, column=3).value == self.MSAName):
                #PeriodYear for unemployment data sheet
                pyUnemp = str(sheet.cell(row=i, column=2).value) + str(sheet.cell(row=i, column=1).value)
                #Find picture that is same quarter as this row
                for root, dirs, files in os.walk(self.picsDir):
                    for fileName in files:
                        #Ensure that we only consider jpgs
                        if (fileName.endswith(".jpg")):
                            pyPic = self.findPicPY(fileName)
                            #The period should never be 0. If it is, this indicates findPicPY() 
                            #is improperly implemented or picture is improperly named
                            if (pyPic == 0):
                                print("Picture " + fileName + " is either indexed or named incorrectly.")
                            #If a period-year data match is found, move to correct directory
                            if (pyUnemp == pyPic):
                                print("Moving picture " + fileName + " to folder " + str(sheet.cell(row=i, column=5).value))
                                self.movePic(fileName, str(sheet.cell(row=i, column=5).value))
                #If the next row isn't for our MSA, stop the loop
                if (sheet.cell(row=i+1, column=3).value != self.MSAName):
                    break
    
    #Move picture to folder whose name is the *1/4-rounded* unemployment
    def movePic(self, fileName, unempRate):
        finalDir = self.rootDir + "/final"
        subDir = finalDir + "/" + unempRate

        #Ensure there is a directory to hold the final arrangement of data
        if ("final" not in os.listdir(self.rootDir)):
            os.mkdir(finalDir)

        existingSubDirs = os.listdir(finalDir)

        #Ensure there is a subdirectory for the unemployment rate corresponding to picture "fileName"
        if (unempRate not in existingSubDirs):
            os.mkdir(subDir)

        #Move image to final location
        os.rename(self.picsDir + "/" + fileName, subDir + "/" + fileName)
        print(fileName + " moved to directory " + subDir + "/" + fileName)

    #Find and concatenate the period and year of the picture
    #This entirely depends on fileName format -- see output of jpgTimeAverager
    def findPicPY(self, fileName):

        period = fileName[46]
        print(period)
        year = fileName[48] + fileName[49] + fileName[50] + fileName[51]
        print(year)

        return str(period) + year
